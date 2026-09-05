"""Generate Dream-Drive MS functionality tracker (one tab per module)."""
from __future__ import annotations

from pathlib import Path

from openpyxl import Workbook
from openpyxl.formatting.rule import FormulaRule
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.worksheet.table import Table, TableStyleInfo
from openpyxl.chart import PieChart, Reference
from openpyxl.chart.label import DataLabelList
from openpyxl.chart.series import DataPoint
from openpyxl.drawing.fill import PatternFillProperties, ColorChoice

OUT = Path(__file__).resolve().parents[1] / "docs" / "Dream-Drive-Functionality-Tracker.xlsx"

STATUSES = ("Done", "In Progress", "Under Review", "Not Started")
HEADERS = [
    "ID",
    "Functionality",
    "Actor",
    "Implementation (what exists today)",
    "Code / API",
    "Status",
    "Priority",
    "Gaps / next work",
]

FILLS = {
    "header": PatternFill("solid", fgColor="1B4332"),
    "Done": PatternFill("solid", fgColor="C6EFCE"),
    "In Progress": PatternFill("solid", fgColor="FFF2CC"),
    "Under Review": PatternFill("solid", fgColor="FCE4D6"),
    "Not Started": PatternFill("solid", fgColor="F4CCCC"),
    "alt": PatternFill("solid", fgColor="F8FAF9"),
    "summary": PatternFill("solid", fgColor="0B3D2E"),
    "card": PatternFill("solid", fgColor="D8F3DC"),
}
FONTS = {
    "header": Font(name="Calibri", bold=True, color="FFFFFF", size=11),
    "title": Font(name="Calibri", bold=True, color="1B4332", size=18),
    "sub": Font(name="Calibri", color="40916C", size=11),
    "body": Font(name="Calibri", size=10),
    "bold": Font(name="Calibri", bold=True, size=10),
    "status": {
        "Done": Font(name="Calibri", bold=True, color="006100", size=10),
        "In Progress": Font(name="Calibri", bold=True, color="9C5700", size=10),
        "Under Review": Font(name="Calibri", bold=True, color="C65911", size=10),
        "Not Started": Font(name="Calibri", bold=True, color="9C0006", size=10),
    },
}
THIN = Border(
    left=Side(style="thin", color="D8E2DC"),
    right=Side(style="thin", color="D8E2DC"),
    top=Side(style="thin", color="D8E2DC"),
    bottom=Side(style="thin", color="D8E2DC"),
)
WRAP = Alignment(wrap_text=True, vertical="top")
LEFT = Alignment(wrap_text=True, vertical="center")

# Each module: (tab, title, service, priority, overall, notes, rows)
# row = (id, functionality, actor, implementation, code, status, priority, gaps)
MODULES = [
    (
        "01 Auth & Security",
        "01 — Auth & security baseline",
        "identity (apps/api)",
        "P0",
        "Done",
        "Firebase + Postgres user sync, OTP (Postgres 5-min TTL), session mint, roles, last-super-admin lock, staff invite, audit+IP, CORS allowlist. Google button shows when NEXT_PUBLIC_GOOGLE_CLIENT_ID is set. Staff city/branch list filters live in module 16.",
        [
            ("01.01", "Firebase ID token verification", "Backend", "Verifies Firebase ID tokens via firebase-admin, with Google tokeninfo fallback. Project ID is enough for ID-token verify.", "apps/api/src/auth.middleware.ts → verifyFirebase", "Done", "P0", "Confirm Firebase authorized domains before prod."),
            ("01.02", "Dev auth bypass (local)", "Backend", "Bearer dev:email upserts a user unless NODE_ENV=production or DEV_AUTH_BYPASS=false.", "auth.middleware.ts resolveUser · session-token.ts allowDevAuthBypass", "Done", "P0", "Must stay off in production."),
            ("01.03", "Auth sync into PostgreSQL", "Customer / Backend", "POST /v1/auth/sync upserts User + CustomerProfile + CUSTOMER role + empty wallet/loyalty. Disabled accounts are rejected.", "POST /v1/auth/sync · identity.service.ts upsertFromIdentity", "Done", "P0", "Gateway leftover service is unused; monolith API does this itself."),
            ("01.04", "Email OTP send", "Customer", "POST /v1/auth/otp/send stores a hashed 5-min code in EmailOtp (Postgres) and emails via notification service. Dev returns devCode.", "POST /v1/auth/otp/send · EmailOtp", "Done", "P0", "Production must not expose devCode (already gated on NODE_ENV)."),
            ("01.05", "Email OTP verify", "Customer", "POST /v1/auth/otp/verify checks hash, upserts user, mints HMAC session token (dd1.*).", "POST /v1/auth/otp/verify · session-token.ts", "Done", "P0", ""),
            ("01.06", "OTP rate limit", "Backend", "3 requests / 15 minutes per IP and per email. 5 verify attempts then lock.", "auth.middleware.ts rateLimitOtp · EmailOtp.windowCount", "Done", "P0", ""),
            ("01.07", "Current user profile", "Customer", "GET /v1/me and PATCH /v1/me (name, phone, address JSON).", "GET/PATCH /v1/me", "Done", "P0", "Phone change does not re-trigger OTP yet."),
            ("01.08", "FCM / device token register", "Mobile / Web", "POST /v1/me/devices stores DeviceToken.", "POST /v1/me/devices", "Done", "P2", "Mobile app does not call this yet."),
            ("01.09", "Admin user search", "Admin", "GET /v1/admin/users?q=&take= for SUPPORT/SALES/CITY_MANAGER/SUPER_ADMIN. Admin /customers search + role editor.", "GET /v1/admin/users · admin /customers", "Done", "P0", ""),
            ("01.10", "Assign staff roles", "Admin", "PATCH /v1/admin/users/:id/roles and POST /v1/admin/users/invite. Last remaining SUPER_ADMIN cannot be removed.", "PATCH .../roles · POST .../invite", "Done", "P0", ""),
            ("01.11", "Disable user", "Admin", "POST /v1/admin/users/:id/disable. Cannot disable self or the last SUPER_ADMIN. Disabled users cannot authenticate.", "POST /v1/admin/users/:id/disable", "Done", "P0", ""),
            ("01.12", "Audit log", "Admin", "GET /v1/admin/audit + admin /audit. Identity writes: sync, login, register, google, otp, roles, invite, disable (with IP).", "GET /v1/admin/audit · AuditLog · apps/admin/app/audit", "Done", "P0", "Refunds/KYC/status audits belong to those modules."),
            ("01.13", "RBAC helpers", "Backend", "requireRoles / requireStaff / assertInternal / assertOwnerOrStaff. CUSTOMER cannot hit /v1/admin. SUPER_ADMIN bypasses role checks.", "apps/api/src/lib/auth.ts · auth.middleware.ts", "Done", "P0", "StaffScope (city/branch) query filters are module 16."),
            ("01.14", "Public vs authed routes", "Backend", "Public: /v1/public/*, OTP, login, register, google, webhooks. Other /v1 requires Bearer. CORS allowlist + security headers.", "auth.middleware.ts PUBLIC · main.ts CORS", "Done", "P0", "Internal routes live at /internal/* (not /v1/internal)."),
            ("01.15", "Web login", "Customer", "/login supports password, email OTP session, register, optional Google, and local dev token. Header Account/Sign in. Sign-out on /account.", "apps/web/src/ms/pages/Login.jsx · AuthContext.jsx", "Done", "P0", "Google button needs NEXT_PUBLIC_GOOGLE_CLIENT_ID."),
            ("01.16", "Admin login", "Admin", "Staff email/password or local mock. Rejects CUSTOMER-only. Sign out in shell.", "apps/admin/app/login/page.jsx · Shell.jsx", "Done", "P0", ""),
            ("01.17", "Google / social login", "Customer", "POST /v1/auth/google verifies Google/Firebase ID token and mints a session. Web GIS button when client id is set.", "POST /v1/auth/google · Login.jsx", "In Progress", "P2", "Set NEXT_PUBLIC_GOOGLE_CLIENT_ID + GOOGLE_OAUTH_CLIENT_ID to enable the button."),
        ],
    ),
    (
        "02 Public CMS",
        "02 — Public website & CMS",
        "platform + apps/web",
        "P1",
        "In Progress",
        "Ported marketing site still runs in Next.js (home, blogs, monsoon bar). New CMS APIs exist for pages, banners, blogs, contact. Admin CMS page is a first version. Testimonials/comments still largely on the old UI.",
        [
            ("02.01", "Marketing home", "Customer", "Ported Home + hero + why-us + fleet teaser + monsoon promo bar (still hardcoded CSS/content).", "apps/web/src/components/home/Home.jsx · MonsoonPromoBar.jsx", "In Progress", "P1", "Drive strips/banners from GET /v1/public/banners."),
            ("02.02", "Public CMS pages", "Customer", "GET /v1/public/pages/:slug returns published CmsPage.", "GET /v1/public/pages/:slug", "Done", "P1", "Web legal pages still static TermsAndConditions, not CMS-driven."),
            ("02.03", "Public banners", "Customer", "GET /v1/public/banners respects active + date window.", "GET /v1/public/banners", "Done", "P1", "Home does not consume this API yet."),
            ("02.04", "Public blogs list/detail", "Customer", "GET /v1/public/blogs and /:slug. Old AllBlogs/Blogspage UI still present.", "GET /v1/public/blogs · apps/web blogs components", "In Progress", "P1", "Point web blogs at platform API instead of Firestore context."),
            ("02.05", "Contact form → CRM lead", "Customer", "POST /v1/public/contact creates Lead + optional activity note.", "POST /v1/public/contact", "Done", "P1", "Old Contact.jsx may still write the previous backend — switch to API."),
            ("02.06", "Admin CMS pages", "Admin", "GET/POST /v1/admin/cms upsert by slug. Admin /cms page exists.", "CRUD /v1/admin/cms · apps/admin/app/cms/page.jsx", "In Progress", "P1", "No unpublish workflow beyond published flag; no media picker."),
            ("02.07", "Admin banners", "Admin", "GET/POST/PATCH /v1/admin/banners.", " /v1/admin/banners", "In Progress", "P1", "No delete endpoint; admin UI may be combined into CMS page."),
            ("02.08", "Testimonials", "Customer / Admin", "Testimonial model in Prisma. Live UI uses old TestimonialContext (ported site).", "schema Testimonial · apps/web Testimonial.jsx", "Not Started", "P1", "Add public + admin APIs and wire home."),
            ("02.09", "Blog comments", "Customer", "BlogComment on schema; public blog include comments. No POST comment API.", "schema BlogComment", "Not Started", "P2", "Auth or moderation flag per spec."),
            ("02.10", "SEO / page metadata", "Admin", "PageMetadata model. Web has SeoDefaults.jsx on old site.", "schema PageMetadata · SeoDefaults.jsx", "In Progress", "P1", "Admin metadata CRUD not on new API."),
            ("02.11", "WhatsApp / enquiry CTAs", "Customer", "Ported WhatsAppPopup + ContactPopup + Numberattach.", "apps/web WhatsAppPopup · ContactPopup", "In Progress", "P1", "Clicks are not logged as CRM leads yet."),
        ],
    ),
    (
        "03 Customer Dashboard",
        "03 — Customer accounts & web dashboard",
        "identity + booking + payment + document + web",
        "P0",
        "In Progress",
        "Logged-in /account lists bookings and KYC status. APIs exist for invoices, wallet, tickets, agreements-by-id. Dedicated screens for invoices/wallet/tickets/agreements are not built yet.",
        [
            ("03.01", "Account home", "Customer", "/account shows name, KYC status, booking list with track links.", "apps/web/src/ms/pages/Account.jsx · GET /v1/me/bookings", "In Progress", "P0", "No profile edit form, addresses UI, or invoices/wallet/tickets links."),
            ("03.02", "My bookings", "Customer", "GET /v1/me/bookings returns history + payments.", "GET /v1/me/bookings", "Done", "P0", "No /account/bookings/[id] detail page (uses /track/:id)."),
            ("03.03", "KYC from account", "Customer", "/account/kyc uploads Aadhaar/PAN/DL/selfie/address, hashed ID numbers, DL expiry, rejection/re-upload UX. GET /v1/me/kyc.", "Kyc.jsx · GET /v1/me/kyc · POST /v1/kyc/uploads", "Done", "P0", "Signed Cloudinary slots need CLOUDINARY_API_SECRET (server multipart works without it)."),
            ("03.04", "Invoices download", "Customer", "GET /v1/me/invoices returns Invoice + lines. No account invoices screen.", "GET /v1/me/invoices", "In Progress", "P0", "Build /account/invoices and PDF download."),
            ("03.05", "Wallet view", "Customer", "GET /v1/me/wallet upserts Wallet and returns txns. No UI.", "GET /v1/me/wallet", "In Progress", "P2", "No credit/debit at checkout (module 18)."),
            ("03.06", "Support tickets", "Customer", "POST /v1/tickets creates ticket + first message. No /account/tickets UI.", "POST /v1/tickets", "In Progress", "P2", "No customer reply thread API beyond create."),
            ("03.07", "Agreements list/download", "Customer", "/account/agreements lists status; authenticated PDF and signed-pdf download. GET /v1/me/documents and GET /v1/me/agreements/:id.", "AccountAgreements.jsx · GET /v1/me/agreements/:id/pdf", "Done", "P0", "Live signed file from Leegality is module 10."),
            ("03.08", "No impersonation", "Admin", "No impersonate API (support is read-only by design).", "docs/modules/03-customer-dashboard.md", "Done", "P0", "Keep it this way."),
        ],
    ),
    (
        "04 Search & Availability",
        "04 — Search, discovery & availability",
        "catalog",
        "P0",
        "Done",
        "Public search, car detail calendar, availability, seasonal pricing, buffer settings, vehicle blocks, HOLD reserve/release/confirm with a serializable lock, and the worker HOLD sweeper.",
        [
            ("04.01", "Public search", "Customer", "GET /v1/public/search filters city, type, seats, fuel, transmission, price, rentalType; sorts featured/price/popularity; returns from-price + available flag.", "GET /v1/public/search · catalog.service.ts search", "Done", "P0", "Unpublished models excluded. Past dates and max rental length rejected."),
            ("04.02", "Car detail by slug", "Customer", "GET /v1/public/cars/:slug with images, pricing, city. Calendar of busy days on CarDetail.", "GET /v1/public/cars/:slug · CarDetail.jsx", "Done", "P0", ""),
            ("04.03", "Availability check", "Customer", "GET /v1/public/cars/:id/availability uses buffer hours, free-vehicle search, and month busyDays.", "GET /v1/public/cars/:id/availability", "Done", "P0", ""),
            ("04.04", "Overlap + buffer + maintenance", "Backend", "vehicleBusy checks blocking booking statuses, AvailabilityBlock, MaintenanceJob. Buffer from CatalogSettings (default 3h).", "catalog.service.ts vehicleBusy · findFreeVehicle", "Done", "P0", "Serializable tx + SELECT FOR UPDATE on reserve."),
            ("04.05", "HOLD reserve / release / confirm", "Backend", "Internal reserve creates HOLD:bookingId block; confirm renames to BOOKING:; release deletes.", "POST /internal/availability/reserve|release|confirm", "Done", "P0", "Requires x-internal-token."),
            ("04.06", "HOLD sweeper", "System", "Worker cron every minute expires unpaid HOLD/AWAITING_PAYMENT and releases availability.", "apps/worker hold-sweeper.job.ts · POST /internal/holds/expire", "Done", "P0", "HOLD_MINUTES env (default 15)."),
            ("04.07", "Admin car models CRUD", "Admin", "GET/POST/PATCH/DELETE /v1/admin/car-models with images, published, featured. Admin /cars page.", " /v1/admin/car-models · apps/admin/app/cars", "Done", "P0", ""),
            ("04.08", "Pricing rules CRUD", "Admin", "GET/POST/PUT/DELETE /v1/admin/pricing-rules (daily/hourly/extraKm/deposit + seasonal startsOn/endsOn).", " /v1/admin/pricing-rules", "Done", "P0", ""),
            ("04.09", "Web search UI", "Customer", "/fleet Search.jsx + /cars redirect. Home search bar. Type filter and price/popularity sort.", "apps/web/src/ms/pages/Search.jsx", "Done", "P0", ""),
            ("04.10", "Seasonal / featured pricing", "Admin", "featured + displayOrder on CarModel. Date-ranged PricingRule.startsOn/endsOn applied at search and quote.", "CatalogSettings · PricingRule", "Done", "P1", ""),
        ],
    ),
    (
        "05 Booking Engine",
        "05 — Booking engine (all rental types)",
        "booking",
        "P0",
        "Done",
        "Quote → HOLD → pay token → KYC/sign (self-drive) → CONFIRMED. All rental types price from rules (hourly chauffeur, daily others, one-way city-pair fee, tour package). Policy cancel + refund, admin create/re-quote, vehicle/driver guards, NO_SHOW sweeper, Razorpay.js checkout, guest track OTP, socket subscribe.",
        [
            ("05.01", "Create quote (price freeze)", "Customer", "POST /v1/quotes prices by rental type (hourly local chauffeur, daily others, airport hourly ≤8h + wait/night, ONE_WAY city-pair fee, tour package price, outstation driver nights). TTL 20 min. GET /v1/quotes/:id hydrates car/branches.", "POST/GET /v1/quotes · booking.service.ts quote", "Done", "P0", ""),
            ("05.02", "Apply offer on quote", "Customer", "POST /v1/quotes/:id/apply-offer discounts from stored grossPaise (not stacking). Floor = max(₹1, 10% of gross).", "POST /v1/quotes/:id/apply-offer", "Done", "P1", ""),
            ("05.03", "Create booking from quote", "Customer", "POST /v1/bookings creates HOLD, copies extras, reserves vehicle in a catalog transaction, moves to AWAITING_PAYMENT or cancels if no car.", "POST /v1/bookings", "Done", "P0", "Quote TTL vs HOLD sweeper both apply."),
            ("05.04", "Get booking + history", "Customer / Admin", "GET /v1/bookings/:id includes history, extras, payments/refunds, KYC, agreements, driver, inspections, vehicle, branches. Owner or staff.", "GET /v1/bookings/:id", "Done", "P0", ""),
            ("05.05", "Cancel booking", "Customer / Admin", "Allowed until HANDOVER. Hours-before-start policy by rental type writes Refund rows (100/50/0%). Releases availability.", "POST /v1/bookings/:id/cancel", "Done", "P0", "Razorpay refund API is module 06."),
            ("05.06", "Lifecycle after payment", "Backend", "TOKEN/BALANCE success → SELF_DRIVE AWAITING_KYC else CONFIRMED. KYC → AWAITING_SIGNATURE. Signed → CONFIRMED. Self-drive handover blocked without KYC+sign. Admin CONFIRMED requires paid token unless comped.", "internal /payment-captured /kyc-approved /agreement-signed", "Done", "P0", "With-driver skips DL KYC/sign by design."),
            ("05.07", "Admin bookings list / patch / status", "Admin", "List filters status/q/dates. PATCH dates re-quotes and re-reserves. Status POST supports comped. Admin /bookings and /bookings/[id] UI.", "/v1/admin/bookings* · apps/admin/app/bookings", "Done", "P0", ""),
            ("05.08", "Assign vehicle", "Admin", "Validates same model, AVAILABLE, insurance expiry, overlap via catalog reserve (row lock).", "POST .../assign-vehicle", "Done", "P0", ""),
            ("05.09", "Assign driver", "Admin", "Active, same city as pickup, not on leave, no overlapping CONFIRMED+ trip. Blocked on SELF_DRIVE.", "POST .../assign-driver", "Done", "P1", ""),
            ("05.10", "Status history", "Backend", "BookingStatusHistory written on create, pay, KYC, sign, cancel, expire, no-show, handover, return. Socket emit on each setStatus.", "BookingStatusHistory", "Done", "P0", ""),
            ("05.11", "Web checkout", "Customer", "/checkout/:quoteId loads quote, login gate, offer, HOLD, mock verify or Razorpay.js, polls payment, success branches self-drive vs chauffeur.", "Checkout.jsx", "Done", "P0", "Live Razorpay still needs keys (module 06)."),
            ("05.12", "Tracking page", "Customer", "/track/:bookingId timeline + driver/vehicle. Logged-out: publicId + phone OTP (matches User.phone).", "Track.jsx · POST /v1/public/bookings/track/*", "Done", "P0", ""),
            ("05.13", "Realtime status", "Customer", "API POSTs /internal/booking-status to socket :4010. Web Track + admin detail subscribe (socket.io CDN) with poll fallback.", "apps/socket + bookingSocket.js", "Done", "P1", "Socket process must be running for live push."),
            ("05.14", "Admin create booking on behalf", "Admin", "POST /v1/admin/bookings with userId or customerEmail; optional extras, notes, flight, comped (waives token).", "POST /v1/admin/bookings", "Done", "P1", ""),
            ("05.15", "NO_SHOW / grace", "Worker", "Worker every 10 min POSTs /internal/bookings/mark-no-show. CONFIRMED/AWAITING_KYC/SIGNATURE past start+2h → NO_SHOW and release.", "NoShowSweeperJob", "Done", "P1", "Grace hours via NO_SHOW_GRACE_HOURS (default 2)."),
        ],
    ),
    (
        "06 Payments",
        "06 — Payments, invoices & deposits",
        "payment",
        "P0",
        "In Progress",
        "Razorpay order + verify + webhook + mock mode, state-wise GST invoices (CGST/SGST vs IGST) with generated PDF, offline pay, Razorpay refund API, deposit auto-release on return, full admin UI (payments/invoices/deposits/recon). Settlement recon runs against Razorpay when keys are set, otherwise mock.",
        [
            ("06.01", "Create payment order (token/balance/deposit)", "Customer", "POST /v1/payments/orders uses booking amount; BALANCE computes remaining due. Mock order if no RAZORPAY_KEY_ID or PAYMENTS_MOCK=true.", "POST /v1/payments/orders", "Done", "P0", ""),
            ("06.02", "Client verify", "Customer", "POST /v1/payments/verify HMAC of order|payment; mock marks SUCCESS.", "POST /v1/payments/verify", "Done", "P0", "Webhook should remain source of truth in prod."),
            ("06.03", "Razorpay webhook", "System", "POST /v1/webhooks/razorpay signature check + idempotent eventId on captured/authorized + refund.processed webhook.", "POST /v1/webhooks/razorpay", "Under Review", "P0", "Test with Razorpay test-mode webhook secret."),
            ("06.04", "Invoice on success", "Customer", "Creates INV-YEAR-##### with CGST/SGST vs IGST by state. GET /v1/me/invoices/:id/pdf returns a generated GST PDF. Account + admin download buttons.", "payment.service.ts afterSuccess · invoice-pdf.ts", "Done", "P1", ""),
            ("06.05", "List my invoices", "Customer", "GET /v1/me/invoices. Customer /account/invoices page with print/download.", "GET /v1/me/invoices · AccountInvoices.jsx", "Done", "P0", ""),
            ("06.06", "Offline cash/UPI", "Admin", "POST /v1/admin/payments/offline marks SUCCESS and runs afterSuccess + audit. Admin UI modal.", "POST /v1/admin/payments/offline", "Done", "P1", "No branch-scoping of who can record."),
            ("06.07", "Refunds", "Admin", "POST /v1/admin/payments/:id/refund calls Razorpay refund API for live payments, writes Refund row. Validates max refundable. Admin UI refund modal.", "POST /v1/admin/payments/:id/refund", "Done", "P1", "Refund webhook updates razorpayRefundId."),
            ("06.08", "Security deposit capture/release", "Admin", "POST deposits/:id/capture|release toggles held/released. Auto-release triggered on fleet return inspection via internal endpoint. Admin UI capture/release buttons.", "POST /v1/admin/deposits/:id/* · internal/deposits/release-by-booking", "Done", "P1", "Razorpay authorize not implemented (uses manual capture)."),
            ("06.09", "Wallet read", "Customer", "GET /v1/me/wallet.", "GET /v1/me/wallet", "In Progress", "P2", "Cannot pay token from wallet."),
            ("06.10", "Admin payments UI", "Admin", "Tabbed admin page: payments (refund + offline), invoices (GST + PDF), deposits (capture/release), recon (run settlement match).", "apps/admin/app/payments/page.jsx", "Done", "P0", ""),
            ("06.11", "Razorpay settlement recon", "Admin", "POST /v1/admin/payments/reconcile pulls Razorpay combined settlement report (3 months) or mock-matches when PAYMENTS_MOCK. Upserts PayoutReconciliation (MATCHED/MISMATCH/UNMATCHED_*). Admin Recon tab.", "POST/GET /v1/admin/payments/reconcile · apps/admin payments Recon tab", "Done", "P2", "Live recon needs RAZORPAY_KEY_ID + secret; mock mode writes MATCHED/OFFLINE rows."),
        ],
    ),
    (
        "07 Subscription",
        "07 — Car subscription & long-term",
        "booking + catalog",
        "P2",
        "In Progress",
        "Data model + POST /v1/subscriptions creates a SUBSCRIPTION booking from a plan. No public plans API, swap, pause, or recurring invoices. After-MVP.",
        [
            ("07.01", "Subscription plan table", "Backend", "SubscriptionPlan (months, price, carModel) in Prisma.", "packages/database/prisma/schema.prisma SubscriptionPlan", "Done", "P2", "No admin CRUD API for plans."),
            ("07.02", "Start subscription", "Customer", "POST /v1/subscriptions quotes SUBSCRIPTION, overwrites amount to plan price, creates booking + Subscription row.", "POST /v1/subscriptions", "In Progress", "P2", "Starts immediately; no handover-aligned start; needs available vehicle."),
            ("07.03", "Public plans listing", "Customer", "GET /v1/public/subscriptions/plans not implemented.", "—", "Not Started", "P2", ""),
            ("07.04", "Swap / early close", "Admin", "POST /v1/admin/subscriptions/:id/swap|close not implemented.", "—", "Not Started", "P2", ""),
            ("07.05", "Monthly invoice / mandate", "Finance", "Spec says invoice-per-month; not built.", "—", "Not Started", "P2", "Razorpay mandate later."),
        ],
    ),
    (
        "08 Airport & Packages",
        "08 — Airport, outstation, one-way & tours",
        "booking + fleet + catalog",
        "P1–P2",
        "Done",
        "Airport wait/night, outstation driver allowance, one-way city-pair fee, and tour packages are quoted and bookable. Admin /packages configures terminals, city pairs, tours, and nightly allowance. Web /packages lists tours. SALES can add extra km/hours on return.",
        [
            ("08.01", "Public tour packages", "Customer", "GET /v1/public/packages and /:slug return published TourPackage + TourDay + city. Web /packages listing and detail.", "GET /v1/public/packages · Packages.jsx", "Done", "P2", ""),
            ("08.02", "Admin packages CRUD", "Admin", "CRUD with days, inclusions, car class, city, itinerary daysDetail. Soft-unpublish via POST delete. Admin /packages tab.", "/v1/admin/packages · apps/admin/app/packages", "Done", "P2", ""),
            ("08.03", "City-pair one-way fee", "Admin / Customer", "GET /v1/public/city-pairs + admin upsert/delete. Quote requires drop city ≠ pickup and a configured pair; fee added to amount.", "/v1/admin/city-pairs · quote ONE_WAY", "Done", "P1", ""),
            ("08.04", "Quote product types", "Customer", "AIRPORT hourly ≤8h + wait/night; OUTSTATION daily + driver nights; ONE_WAY daily + pair fee; TOUR_PACKAGE fixed package price (car class must match).", "POST /v1/quotes rentalType", "Done", "P1", "Wait is a single slab after free minutes, not multi-slab."),
            ("08.05", "Flight number & terminal", "Customer", "Car detail collects terminal, optional flight, wait minutes. Stored on quote payload and Booking.", "CarDetail.jsx · Booking.flightNumber/terminalId", "Done", "P1", ""),
            ("08.06", "Airport terminals & wait/night", "Admin", "CRUD /v1/admin/airports + GET /v1/public/airports. Free wait minutes, paise/min, night window (IST) and surcharge.", "AirportTerminal · fleet airports", "Done", "P1", ""),
            ("08.07", "Outstation driver allowance", "Backend", "CatalogSettings.driverAllowancePerNightPaise × IST midnights crossed, auto-added as a quote extra.", "priceWindow OUTSTATION", "Done", "P1", ""),
            ("08.08", "Web packages page", "Customer", "/packages and /packages/:slug. Header Tours link. Book → /fleet with packageId + car class.", "apps/web/src/ms/pages/Packages.jsx", "Done", "P2", ""),
            ("08.09", "Extra km / hours on return", "Admin", "POST /v1/admin/bookings/:id/extras. SALES may override before return; others only on HANDOVER+.", "addBookingExtras · booking detail UI", "Done", "P1", "Does not auto-compute extra km from odometer (handover module)."),
        ],
    ),
    (
        "09 KYC & Documents",
        "09 — Legal KYC & documents",
        "document",
        "P0",
        "Under Review",
        "Customer upload (MIME/size) + hashed Aadhaar/PAN last4 + DL vs drop-off + 12-month reuse + admin review/reject/re-upload. Zoho webhook HMAC + attachment pull need staging secrets. Signed Cloudinary slots need API_SECRET; unsigned server upload works today.",
        [
            ("09.01", "Submit KYC case", "Customer", "POST /v1/kyc/submit with docs + optional Aadhaar/PAN/DL expiry. Validates kinds, required self-drive set (DL + Aadhaar/address + selfie), updates REJECTED cases in place.", "POST /v1/kyc/submit · document.service.ts submitKyc", "Done", "P0", ""),
            ("09.02", "Signed upload slots", "Customer", "POST /v1/kyc/uploads issues a Cloudinary signed slot when API_KEY+SECRET are set; otherwise returns mode=server. Multipart file is validated (pdf/jpg/png/webp, 8 MB) and uploaded via the API. Production rejects arbitrary URLs.", "POST /v1/kyc/uploads · lib/cloudinary.ts issueCloudinarySlot", "Under Review", "P0", "Set CLOUDINARY_API_KEY + CLOUDINARY_API_SECRET for direct-to-Cloudinary signed uploads. Unsigned preset still used for server-side uploads."),
            ("09.03", "My KYC", "Customer", "GET /v1/me/kyc returns cases, documents, last4, validUntil, reusable flag. Hashes never leave the API.", "GET /v1/me/kyc · GET /v1/me/documents", "Done", "P0", ""),
            ("09.04", "Admin KYC queue + decision", "Admin", "GET /v1/admin/kyc with status filter, DL expiry flags. POST review → UNDER_REVIEW. Decision APPROVED/REJECTED (reason required). Request re-upload per doc. SUPPORT + SUPER_ADMIN decide; SALES view-only. Admin /kyc has doc viewer.", " /v1/admin/kyc · apps/admin/app/kyc", "Done", "P0", "No KYC_OFFICER role in RoleName — SUPPORT is the designated KYC role."),
            ("09.05", "Zoho Forms webhook", "System", "POST /v1/webhooks/zoho-form verifies ZOHO_WEBHOOK_SECRET (header or HMAC) in production, upserts user, matches booking by email+dates+car, stores hashed Aadhaar/PAN, strips raw IDs from ZohoSubmission.raw. Downloads attachments to Cloudinary when Zoho OAuth env is set.", "POST /v1/webhooks/zoho-form · lib/zoho.ts", "Under Review", "P0", "Needs ZOHO_WEBHOOK_SECRET in prod. Attachment pull needs ZOHO_CLIENT_ID/SECRET/REFRESH_TOKEN."),
            ("09.06", "Aadhaar / PAN storage policy", "Backend", "Raw numbers hashed with KYC_HASH_PEPPER (SHA-256); only last4 stored. Never written to audit logs or Zoho raw JSON.", "lib/kyc.ts identityStamp · KycCase.aadhaarHash", "Done", "P0", "Set KYC_HASH_PEPPER in production (falls back to SESSION_SECRET)."),
            ("09.07", "DL expiry vs drop-off", "Backend", "DL expiry required on submit; must be after booking.endsAt. Handover also blocks if DL misses drop-off. Admin queue flags expired / expiring / misses-drop-off.", "submitKyc assertDlCoversDropOff · fleet handover", "Done", "P0", ""),
            ("09.08", "Reusable KYC 12 months", "Backend", "On approve, validUntil = min(12 months, DL expiry). Payment capture calls /internal/kyc/apply-reusable to clone an approved case onto the new booking and skip AWAITING_KYC.", "applyReusable · kycValidUntil", "Done", "P1", ""),
            ("09.09", "Web KYC UI", "Customer", "Per-doc file pickers (Aadhaar, PAN, DL, selfie, address), hashed ID fields, DL date, rejection reason, re-upload highlighting, reusable-approved state.", "apps/web/src/ms/pages/account/AccountKyc.jsx", "Done", "P0", ""),
        ],
    ),
    (
        "10 Agreement & eSign",
        "10 — Rental agreement & Leegality",
        "document",
        "P0",
        "Under Review",
        "PDF generate (vehicle, tariff, jurisdiction) + mock/live Leegality send + webhook HMAC/IP + customer PDF download + template CRUD by city/product + void/re-issue. Live eSign still needs a published Leegality workflow and staging credentials.",
        [
            ("10.01", "Generate agreement", "Admin", "POST /v1/agreements/generate builds a real PDF (parties, vehicle, dates, tariff, deposit, city jurisdiction) from AgreementTemplate HTML. Self-drive requires APPROVED KYC. Reuses an open DRAFT.", "POST /v1/agreements/generate · agreement-pdf.ts · GET /v1/me/agreements/:id/pdf", "Done", "P0", "PDF is Helvetica (same as invoices), not PDFKit. Legal clause copy still needs counsel review."),
            ("10.02", "Send to Leegality", "Admin", "POST /v1/agreements/:id/send-leegality posts v3 /sign/request with profileId + base64 PDF + invitee email matching the customer. Mocks envelope when LEEGALITY_API_KEY or LEEGALITY_PROFILE_ID is missing.", "POST /v1/agreements/:id/send-leegality · leegality.ts createLeegalityRequest", "Under Review", "P0", "BLOCKER: published Leegality workflow (LEEGALITY_PROFILE_ID) + API key. Invite emails/SMS are sent by Leegality from the workflow, not by us. Staging review still required."),
            ("10.03", "Leegality webhook", "System", "POST /v1/webhooks/leegality verifies HMAC-SHA1 mac (documentId + LEEGALITY_PRIVATE_SALT) and optional IP allowlist. Completed → fetch signed file, store SignedArtifact, advance booking.", "POST /v1/webhooks/leegality · leegality.ts verifyLeegalityWebhook", "Under Review", "P0", "BLOCKER: set LEEGALITY_PRIVATE_SALT (required in production). Webhook URL must be configured on the Leegality workflow (cannot set via send API). Set LEEGALITY_IP_ALLOWLIST once vendor IPs are known."),
            ("10.04", "Admin wet-ink / waiver", "Admin", "POST /v1/admin/agreements/:id/mark-signed (SUPPORT/SUPER_ADMIN) sets WAIVED, optional wet-ink URL, advances booking like a signed envelope.", "POST /v1/admin/agreements/:id/mark-signed", "Done", "P0", ""),
            ("10.05", "Customer download", "Customer", "GET /v1/me/agreements/:id plus authenticated PDF and signed-pdf streams. Account Agreements and booking detail download with the session token.", "GET /v1/me/agreements/:id · /pdf · /signed-pdf", "Done", "P0", "Live signed bytes come from Leegality Details API after webhook; mock/waiver stamps the generated PDF."),
            ("10.06", "Templates by city/product", "Admin", "CRUD on AgreementTemplate (name, html, cityId, rentalType, active). Generate picks city+product, then city, then product, then default HTML.", "GET/POST/PATCH/DELETE /v1/admin/agreement-templates · apps/admin/app/agreements", "Done", "P1", ""),
            ("10.07", "Void and re-issue", "Admin", "POST /v1/admin/agreements/:id/void (SUPER_ADMIN) sets VOID + reason; optional reissue generates a replacement DRAFT and links supersededById.", "POST /v1/admin/agreements/:id/void", "Done", "P1", "BLOCKER: void does not recall a live Leegality envelope — cancel that invite in the Leegality dashboard."),
        ],
    ),
    (
        "11 Fleet",
        "11 — Vehicle & fleet management",
        "fleet",
        "P0",
        "Done",
        "Physical vehicles are distinct from CarModel. CRUD, queued or immediate branch transfer, RC/insurance/PUC files (local upload if no Cloudinary), 30-day expiry emails, partner-contract check, and insurance-covers-booking on assignment.",
        [
            ("11.01", "Public cities", "Customer", "GET /v1/public/cities for search.", "GET /v1/public/cities", "Done", "P0", ""),
            ("11.02", "City CRUD", "Admin", "GET/POST/PATCH/DELETE /v1/admin/cities. Vehicles page Cities & branches tab can create cities.", " /v1/admin/cities · apps/admin/app/vehicles", "Done", "P0", "Full city editor (edit/delete) still SUPER_ADMIN/CITY_MANAGER API; tab is create+list."),
            ("11.03", "Branch CRUD", "Admin", "GET/POST/PATCH/DELETE /v1/admin/branches. Vehicles page can create branches.", " /v1/admin/branches · apps/admin/app/vehicles", "Done", "P0", "Edit/delete remain API-only."),
            ("11.04", "Vehicle CRUD", "Admin", "Full CRUD with model/branch/partner/status/odometer. Unique registration. BRANCH_MANAGER scoped when StaffScope is set. PUT /v1/admin/staff/:id/scope on Customers page.", " /v1/admin/vehicles · PUT /v1/admin/staff/:id/scope", "Done", "P0", ""),
            ("11.05", "Move vehicle between branches", "Admin", "POST transfer-branch immediate (default) or queued job. Pending jobs listed; complete/cancel. Queued transfer blocks AVAILABLE cars until complete.", "POST /v1/admin/vehicles/:id/transfer-branch · /v1/admin/vehicle-transfers", "Done", "P1", ""),
            ("11.06", "Document expiry alerts", "Admin", "GET expiries + POST /v1/admin/vehicles/expiries/notify emails FLEET_OPS/CITY_MANAGER. Worker cron 08:00. alertedAt avoids daily spam.", "GET expiries · POST expiries/notify · VehicleExpiryJob", "Done", "P1", "Needs GMAIL_USER for real mail; otherwise console mock."),
            ("11.07", "RC / insurance / PUC files", "Admin", "Document CRUD. Local disk upload when Cloudinary is unset (/v1/files/...). POST backfill-documents fills missing RC/insurance/PUC/permit on existing cars. New vehicles get placeholders automatically.", "VehicleDocument · /v1/uploads · backfill-documents", "Done", "P1", ""),
            ("11.08", "Partner-owned vehicles", "Admin", "ownerType/partnerId on create/update. Active partner + covering PartnerContract required. POST /v1/admin/partners/:id/contracts from the Vehicles geo tab.", "createVehicle · assertPartnerContractActive", "Done", "P2", "Full partner settlement UI is still module 15."),
        ],
    ),
    (
        "12 Maintenance",
        "12 — Maintenance & workshop",
        "fleet",
        "P2",
        "In Progress",
        "Creating a maintenance job writes an availability block so search will not sell that car. Complete-job, parts, workshop CRUD, and admin UI are not built. After-MVP.",
        [
            ("12.01", "Create maintenance job + block", "Admin", "POST /v1/admin/maintenance-jobs creates job and AvailabilityBlock reason MAINT:id.", "POST /v1/admin/maintenance-jobs", "Done", "P2", "Does not stop overlap with ONGOING booking."),
            ("12.02", "List / delete jobs", "Admin", "GET + DELETE. Delete may leave the availability block behind.", "GET/DELETE /v1/admin/maintenance-jobs", "In Progress", "P2", "Delete should release MAINT: block; PATCH missing."),
            ("12.03", "Complete job → AVAILABLE", "Admin", "POST .../complete not implemented.", "—", "Not Started", "P2", "Require odometer + cost."),
            ("12.04", "Parts / labour", "Admin", "MaintenancePart model unused in API.", "schema MaintenancePart", "Not Started", "P2", ""),
            ("12.05", "Workshop master", "Admin", "Workshop model, no CRUD.", "schema Workshop", "Not Started", "P2", ""),
            ("12.06", "Admin maintenance screen", "Admin", "Not in admin Shell nav.", "apps/admin/components/Shell.jsx", "Not Started", "P2", ""),
        ],
    ),
    (
        "13 Handover & Return",
        "13 — Handover & return inspection",
        "fleet",
        "P1",
        "Done",
        "Handover from CONFIRMED (self-drive needs KYC + signed agreement). Return checks odometer ≥ handover and matching fuel unit. Damages raise PENALTY invoices. Deposit + partner ledger wait until the return inspection is CLOSED with no open damages. Admin booking detail + /inspections. Customer sees damage lines on booking and invoice. Phone checklist ack is later. Delivery-exec app is out of scope.",
        [
            ("13.01", "Handover inspection", "Admin", "POST /v1/admin/bookings/:id/handover from CONFIRMED: Inspection HANDOVER + photos + accessories + signature; booking HANDOVER→ONGOING; vehicle ON_TRIP. Self-drive requires approved KYC + signed/waived agreement.", "POST /v1/admin/bookings/:id/handover · inspection.service.ts", "Done", "P1", ""),
            ("13.02", "Return inspection", "Admin", "POST .../return from ONGOING: odometer ≥ handover, same fuel unit, RETURN inspection. No damages → COMPLETED + deposit release. With damages → RETURN_PENDING until close.", "POST /v1/admin/bookings/:id/return", "Done", "P1", ""),
            ("13.03", "Damage line items", "Admin", "Damages on return (or later) create DamageCharge + BookingExtra + Payment PENALTY + GST invoice. Partner trip-complete waits until return CLOSED.", "returnVehicle damages[] · raisePenalty", "Done", "P1", ""),
            ("13.04", "Dedicated damage API", "Admin", "POST /v1/admin/inspections/:id/damages adds lines on an open return. Waive + close endpoints included.", "POST /v1/admin/inspections/:id/damages · /close · /damages/:id/waive", "Done", "P1", ""),
            ("13.05", "Deposit release on closed return", "Finance", "internal/deposits/release-by-booking runs only when the RETURN inspection is CLOSED and no OPEN damages.", "depositReleaseByBooking · POST .../inspections/:id/close", "Done", "P1", ""),
            ("13.06", "Customer acknowledge checklist", "Customer", "Out of MVP; not built. Spec marks optional later.", "—", "Not Started", "P2", "Phone ack of handover checklist."),
            ("13.07", "Admin handover UI", "Admin", "Booking detail: handover (odo, fuel, accessories, photos, signature) and return (+ damages). /inspections list with close action.", "apps/admin/app/bookings/[id] · app/inspections", "Done", "P1", ""),
        ],
    ),
    (
        "14 Drivers",
        "14 — Driver & chauffeur management",
        "fleet",
        "P1",
        "Done",
        "Chauffeur roster with DL/badge docs, leave, and availability. Assign to with-driver bookings (same city, DL through trip end, one ONGOING). CITY_MANAGER can override city. Customers see name/phone after assignment on track and account booking. No driver marketplace.",
        [
            ("14.01", "Driver CRUD", "Admin", "GET/POST/PATCH/DELETE /v1/admin/drivers scoped by branch/city. GET :id includes docs, leave, current assignment. Admin /drivers roster form.", "CRUD /v1/admin/drivers · apps/admin/app/drivers", "Done", "P1", ""),
            ("14.02", "Assign to booking", "Admin", "POST assign-driver: active, not on leave, same city as pickup (CITY_MANAGER/SUPER_ADMIN overrideCity), DL valid through endsAt. Booking GET + admin/customer UIs show name/phone.", "POST /v1/admin/bookings/:id/assign-driver", "Done", "P1", ""),
            ("14.03", "Driver documents", "Admin", "POST/PATCH/DELETE /v1/admin/drivers/:id/documents (DL, BADGE, ID, POLICE_VERIFICATION). DL requires expiry; assign blocks if DL missing or expires before trip end.", "POST .../drivers/:id/documents · dlCovers", "Done", "P1", ""),
            ("14.04", "Leave / availability", "Admin", "POST/PATCH/DELETE leave. GET /v1/admin/drivers/availability?from&to returns available/busy + reasons. Admin Availability tab + booking assign dropdown.", "GET /v1/admin/drivers/availability · DriverLeave", "Done", "P1", ""),
            ("14.05", "One ONGOING assignment", "Backend", "Assign blocked if driver already has HANDOVER/ONGOING/RETURN_PENDING, or overlapping CONFIRMED+ dates.", "assignDriver DRIVER_ON_TRIP + ACTIVE_TRIP clash", "Done", "P1", ""),
            ("14.06", "Customer sees driver after assign", "Customer", "Name + phone on /track/:id and /account/bookings/:id after assignment. Not shown at search time.", "Track.jsx · AccountBookingDetail.jsx", "Done", "P1", "No customer-facing driver marketplace (by design)."),
        ],
    ),
    (
        "15 Partners",
        "15 — Partners, commission & settlement",
        "partner",
        "P2",
        "In Progress",
        "Admin-only partner CRUD, commission rules, ledger on trip complete, settlement generate + mark-paid. No partner portal (confirmed exclusion). Open-damage hold and frozen-at-booking commission version still missing.",
        [
            ("15.01", "Partner CRUD", "Admin", "GET/POST/PATCH/DELETE /v1/admin/partners. Admin /partners page.", " /v1/admin/partners", "Done", "P2", ""),
            ("15.02", "Commission rules", "Admin", "PUT /v1/admin/partners/:id/commission-rules replaces rules (% bps + flat).", "PUT .../commission-rules", "Done", "P2", "Not snapshotted onto the booking at HOLD time."),
            ("15.03", "Ledger on trip complete", "System", "Return inspection calls /internal/ledger/trip-complete → TRIP_EARNING + COMMISSION.", "POST /internal/ledger/trip-complete", "Done", "P2", "Skipped if vehicle has no partnerId."),
            ("15.04", "View ledger", "Admin", "GET /v1/admin/partners/:id/ledger.", "GET .../ledger", "Done", "P2", ""),
            ("15.05", "Generate settlement", "Admin", "POST /v1/admin/settlements/generate sums ledger in period; unique partner+period.", "POST /v1/admin/settlements/generate", "Done", "P2", "Does not exclude open DamageCharge."),
            ("15.06", "Mark paid (UTR)", "Admin", "POST /v1/admin/settlements/:id/mark-paid.", "POST .../mark-paid", "In Progress", "P2", "Bank account required-before-pay not enforced."),
            ("15.07", "Partner self-service portal", "Partner", "Explicitly out of scope — admin manages partners.", "README Out of scope", "Not Started", "—", "Do not build."),
        ],
    ),
    (
        "16 Admin & Multi-city",
        "16 — Admin console, multi-city & branch ops",
        "fleet + platform + apps/admin",
        "P0–P1",
        "In Progress",
        "Ops console (port 3001) has a city/branch switcher, dedicated /cities and /staff pages, and dashboard KPIs (handovers, overdue returns, KYC, signatures, failed payments, workshop). Bookings, vehicles, drivers, KYC, agreements, payments, tickets, and maintenance honor StaffScope plus the switcher. Legacy apps/web /admin is still present.",
        [
            ("16.01", "Admin app shell", "Admin", "Next.js apps/admin with sidebar, token gate, city/branch switcher, /cities and /staff routes.", "apps/admin/components/Shell.jsx · /cities · /staff", "Done", "P0", "Inspections stay on booking detail. Notification-template admin page is still missing."),
            ("16.02", "Dashboard KPIs", "Admin", "GET /v1/admin/dashboard scoped by location: bookings, pending KYC, free vehicles, revenue, today’s handovers, overdue returns, pending signatures, failed payments, workshop, byStatus. Optional from/to.", "GET /v1/admin/dashboard · app/page.jsx", "Done", "P1", ""),
            ("16.03", "City + branch on every entity", "Backend", "City/Branch models; vehicles/drivers/car models/bookings reference them.", "schema City Branch", "Done", "P0", ""),
            ("16.04", "StaffScope assign", "Admin", "PUT /v1/admin/staff/:id/scope. Dedicated /staff page plus Customers invite.", "PUT /v1/admin/staff/:id/scope · app/staff", "Done", "P1", ""),
            ("16.05", "Scope enforcement", "Backend", "x-ops-city-id / x-ops-branch-id overlay for SUPER_ADMIN and CITY_MANAGER. Bookings, vehicles, drivers, KYC, agreements, payments, tickets, maintenance, dashboard filter by pickup branch / vehicle branch.", "vehicleScopeWhere · bookingScopeWhere · AuthMiddleware", "Done", "P1", ""),
            ("16.06", "Legacy web admin", "Admin", "Ported AdminLayout still at /admin on the public site (blogs, form entries, users, rides).", "apps/web Admin/*", "In Progress", "P1", "Retire Firestore admin after cutover."),
        ],
    ),
    (
        "17 Finance & Reports",
        "17 — Finance, accounting & reports",
        "platform + payment",
        "P2",
        "In Progress",
        "Basic revenue + GST report endpoints and an admin reports page. No CSV export, deposit/partner reports, or Razorpay recon queue. GST is a flat 18% on invoice create.",
        [
            ("17.01", "Revenue report", "Admin", "GET /v1/admin/reports/:kind — non-gst kinds return SUCCESS payments + totalPaise.", "GET /v1/admin/reports/revenue", "In Progress", "P2", "No date-range or city filter; take 500."),
            ("17.02", "GST report", "Admin", "kind=gst sums invoice amount + gstPaise.", "GET /v1/admin/reports/gst", "In Progress", "P2", "Not a real GSTR pack."),
            ("17.03", "Bookings / deposits / partners reports", "Admin", "Documented routes not implemented as distinct kinds.", "docs/api-routes.md", "Not Started", "P2", ""),
            ("17.04", "CSV export + audit", "Admin", "Not implemented.", "—", "Not Started", "P1", "P1 in sprint notes for basic CSV."),
            ("17.05", "Admin reports UI", "Admin", "apps/admin/app/reports/page.jsx.", "apps/admin/app/reports", "In Progress", "P2", ""),
            ("17.06", "Razorpay vs internal mismatch", "Finance", "Not built.", "—", "Not Started", "P2", ""),
        ],
    ),
    (
        "18 Offers & Wallet",
        "18 — Offers, loyalty, referral & wallet",
        "platform + payment + booking",
        "P1–P2",
        "In Progress",
        "Coupon apply on quote is real (window, max redemptions, one per user). Admin can create offers. Wallet/loyalty rows are created at signup but earn/redeem/referral are not. P1 = coupons; P2 = wallet/loyalty.",
        [
            ("18.01", "Admin offers CRUD (create/list)", "Admin", "GET/POST /v1/admin/offers (PERCENT/FLAT, dates, maxRedemptions). Admin /offers page.", " /v1/admin/offers", "In Progress", "P1", "No PATCH freeze/disable; no city/product/min-days."),
            ("18.02", "Apply coupon on quote", "Customer", "Checkout can POST apply-offer; booking stores offerId + redemption.", "POST /v1/quotes/:id/apply-offer", "Done", "P1", "No floor price."),
            ("18.03", "Wallet ledger", "Customer", "Wallet + WalletTxn models; GET wallet. Signup creates zero balance.", "GET /v1/me/wallet · identity upsert", "In Progress", "P2", "No credit/debit/adjust API; cannot pay from wallet."),
            ("18.04", "Loyalty points", "Customer", "LoyaltyAccount created at signup. No earn on COMPLETED, no /v1/me/loyalty.", "schema LoyaltyAccount", "Not Started", "P2", ""),
            ("18.05", "Referral codes", "Customer", "Referral model. No POST /v1/me/referrals.", "schema Referral", "Not Started", "P2", "Credit after referee COMPLETED."),
        ],
    ),
    (
        "19 Reviews & Support",
        "19 — Reviews & customer support desk",
        "platform",
        "P2",
        "In Progress",
        "Customers can create a review (unpublished) and a ticket. Admin can list both. No moderate/publish, public car reviews, ticket assign/SLA, or customer ticket UI.",
        [
            ("19.01", "Create review", "Customer", "POST /v1/reviews stored published=false.", "POST /v1/reviews", "In Progress", "P2", "No completed-booking or one-per-booking check."),
            ("19.02", "Admin review list", "Admin", "GET /v1/admin/reviews.", "GET /v1/admin/reviews", "In Progress", "P2", "No publish/hide PATCH; no admin page in Shell."),
            ("19.03", "Public reviews on car", "Customer", "GET /v1/public/cars/:id/reviews missing.", "—", "Not Started", "P2", ""),
            ("19.04", "Open support ticket", "Customer", "POST /v1/tickets with subject/body/bookingId.", "POST /v1/tickets", "Done", "P2", "No /account/tickets UI."),
            ("19.05", "Admin ticket queue", "Admin", "GET /v1/admin/tickets includes messages.", "GET /v1/admin/tickets", "In Progress", "P2", "No PATCH assign/status/internal notes/SLA; no Shell link."),
        ],
    ),
    (
        "20 Notifications",
        "20 — Notification suite",
        "notification + worker",
        "P0",
        "Under Review",
        "Internal notify + templates + logs + retry cron. Gmail OAuth/app-password when env is set; otherwise console mock. OTP and booking_confirmed are wired. SMS/WhatsApp later. Staging review once Gmail env is present.",
        [
            ("20.01", "Internal send", "System", "POST /internal/notify renders template {{vars}} and emails or mocks.", "POST /internal/notify · notify.service.ts", "Done", "P0", "HTML is wrapped in <pre>."),
            ("20.02", "Gmail delivery", "System", "nodemailer Gmail OAuth2 or app password if GMAIL_USER set.", "notify.service.ts deliver", "Under Review", "P0", "Confirm production Gmail OAuth (ported from live backend)."),
            ("20.03", "OTP email", "Customer", "identity sendOtp → template otp. Failures swallowed so login still works in dev.", "POST /v1/auth/otp/send", "In Progress", "P0", "Dev still returns devCode."),
            ("20.04", "Booking confirmation email", "Customer", "Sent on with-driver CONFIRMED and on agreement-signed.", "booking.service.ts paymentCaptured / agreementSigned", "In Progress", "P0", "Port current HTML template; self-drive confirm only after sign."),
            ("20.05", "Templates admin", "Admin", "GET list + PUT /v1/admin/notification-templates/:key.", " /v1/admin/notification-templates", "Done", "P1", "No admin UI."),
            ("20.06", "Delivery log", "Admin", "GET /v1/admin/notifications last 200.", "GET /v1/admin/notifications", "Done", "P1", "PII masking not done."),
            ("20.07", "Retry failed", "System", "Worker every 5 min → /internal/notify/retry. Retry currently re-sends template key as subject/body.", "NotifyRetryJob", "In Progress", "P1", "Store original body; 3x backoff."),
            ("20.08", "SMS / WhatsApp", "Customer", "Same template table, channel later.", "—", "Not Started", "P2", ""),
        ],
    ),
    (
        "21 CRM & Leads",
        "21 — Leads, enquiries & sales CRM",
        "platform",
        "P1",
        "In Progress",
        "Contact and public lead APIs + admin pipeline (NEW→LOST) + notes. Convert-to-booking and 7-day dedupe are not built. Admin /leads page exists.",
        [
            ("21.01", "Public lead capture", "Customer", "POST /v1/public/leads and contact→lead.", "POST /v1/public/leads · POST /v1/public/contact", "Done", "P1", "Phone-or-email required is not strictly enforced (name only)."),
            ("21.02", "Admin lead list", "Admin", "GET /v1/admin/leads with activities. Admin /leads.", "GET /v1/admin/leads", "Done", "P1", ""),
            ("21.03", "Pipeline status", "Admin", "PATCH /v1/admin/leads/:id status NEW|CONTACTED|QUALIFIED|BOOKED|LOST.", "PATCH /v1/admin/leads/:id", "Done", "P1", "No owner assign / reminders."),
            ("21.04", "Lead notes", "Admin", "POST /v1/admin/leads/:id/notes.", "POST .../notes", "Done", "P1", ""),
            ("21.05", "Convert lead → quote/booking", "Admin", "POST /v1/admin/leads/:id/convert missing.", "—", "Not Started", "P1", "Prefill city + dates."),
            ("21.06", "Dedupe 7 days", "Backend", "Not implemented.", "—", "Not Started", "P1", "Match phone+email."),
            ("21.07", "WhatsApp click as source", "Customer", "Popup exists; not tracked as lead.", "WhatsAppPopup.jsx", "Not Started", "P1", ""),
        ],
    ),
    (
        "22 Mobile",
        "22 — Customer mobile app (Expo)",
        "apps/mobile",
        "P2",
        "In Progress",
        "Expo package with a single-screen prototype: dev sign-in, public car list, my bookings. No OTP, checkout, Razorpay, KYC, tracking, tickets, or store build. Spec: do not invest until web checkout is stable.",
        [
            ("22.01", "Expo app shell", "Customer", "App.tsx SafeArea list + dev:email login + GET search + GET my bookings.", "apps/mobile/App.tsx", "In Progress", "P2", "Not a navigation stack / screens 1–13."),
            ("22.02", "Login / OTP", "Customer", "Uses Bearer dev:email only.", "App.tsx", "Not Started", "P2", "Firebase Auth + OTP."),
            ("22.03", "Search / detail / checkout / pay", "Customer", "Cars listed read-only; no quote/checkout/Razorpay.", "—", "Not Started", "P2", "Reuse /v1/*."),
            ("22.04", "KYC + Leegality + track", "Customer", "Not built.", "—", "Not Started", "P2", ""),
            ("22.05", "Push device token", "Customer", "API POST /v1/me/devices ready; app unused.", "POST /v1/me/devices", "Not Started", "P2", "FCM."),
            ("22.06", "Store release (Play/App Store)", "Ops", "No EAS config / force-update.", "apps/mobile/package.json", "Not Started", "P2", "After web path is stable."),
        ],
    ),
]


FE_HEADERS = [
    "ID",
    "Page",
    "Route",
    "How it should function",
    "What exists today + APIs",
    "Status",
    "Priority",
    "Gaps / next work",
]

# Frontend developer tasks: (tab, title, app, priority, overall, notes, rows)
# row = (id, page, route, how it should work, current + APIs, status, priority, gaps)
FE_MODULES = [
    (
        "FE Web Public",
        "Frontend — public website (apps/web)",
        "apps/web :3000",
        "P0–P1",
        "In Progress",
        "Customer-facing marketing + booking path. New MS pages live under /fleet, /login, /checkout, /account, /track. Ported CRA pages still cover home, blogs, contact. Frontend owns UI, forms, Razorpay checkout widget, and calling NEXT_PUBLIC_API_URL (port 4000). No Prisma/Firebase Admin in the browser.",
        [
            (
                "FE.W.01",
                "Home",
                "/",
                "Hero with city + dates search CTA that goes to /fleet. Show CMS banners, featured cars, how-it-works, testimonials, blogs teaser, WhatsApp/enquiry CTAs. Promo strips (sale/monsoon) come from GET /v1/public/banners, not hardcoded CSS.",
                "Ported Home.jsx + Hero + MonsoonPromoBar (hardcoded). Does not call banners API.",
                "In Progress",
                "P1",
                "Wire banners/CMS; search form should land on /fleet with query params.",
            ),
            (
                "FE.W.02",
                "Fleet / search",
                "/fleet  (also /cars)",
                "Filters: city, from/to, rental type (self-drive vs chauffeur), seats, fuel, transmission, price. Results show image, from-price, available flag. Unavailable cars greyed. Sort by price. Clicking a card opens car detail with the same dates.",
                "Search.jsx on /fleet calls GET /v1/public/cities + /v1/public/search (city/dates/SELF_DRIVE only). /cars is still the old FleetCarousel.",
                "In Progress",
                "P0",
                "Add remaining filters; unify /cars vs /fleet; pass dates into detail.",
            ),
            (
                "FE.W.03",
                "Car detail + quote",
                "/cars/:slug",
                "Gallery, specs, city, from-price. Date/time + rental-type picker. Show blocked dates / availability. CTA creates a quote then navigates to /checkout/:quoteId. Guest can browse; checkout requires login.",
                "CarDetail.jsx exists. GET /v1/public/cars/:slug and availability. Calendar of blocked dates is thin.",
                "In Progress",
                "P0",
                "POST /v1/quotes from this page; blocked-date calendar; pickup/drop branch.",
            ),
            (
                "FE.W.04",
                "Login",
                "/login",
                "Email/password (Firebase) and email OTP. On success: POST /v1/auth/sync, store Bearer token, redirect to /account or intended checkout URL. Show errors and OTP resend cooldown.",
                "Login.jsx is a local demo: loginDev(email) only. AuthContext.jsx holds the token.",
                "In Progress",
                "P0",
                "Real Firebase Auth + OTP send/verify UI. Google social login later.",
            ),
            (
                "FE.W.05",
                "Checkout (quote review)",
                "/checkout/:quoteId",
                "Show frozen price, dates, car, deposit, offer-code field. Confirm creates booking (HOLD) then starts token payment. If quote expired, send back to car detail. Login gate if anonymous.",
                "Checkout.jsx loads GET /v1/quotes/:id, offer, login gate, HOLD, mock or Razorpay.js, polls payment.",
                "Done",
                "P0",
                "Live Razorpay needs keys (module 06).",
            ),
            (
                "FE.W.06",
                "Pay (Razorpay)",
                "/checkout/:quoteId",
                "Open Razorpay Checkout with orderId, keyId, amount from POST /v1/payments/orders. On success call POST /v1/payments/verify, then poll GET /v1/payments/:id until SUCCESS (webhook is source of truth). Handle failure/cancel.",
                "Razorpay Checkout.js opens when order.mock is false; verify + poll GET /v1/payments/:id. Mock path still verifies immediately.",
                "Done",
                "P0",
                "Needs RAZORPAY_KEY_ID in API env for the live widget.",
            ),
            (
                "FE.W.07",
                "Booking success",
                "/checkout/success",
                "Show public booking id, amount, next steps: KYC (self-drive) or wait for confirm (with-driver). Links to /account/kyc and /track/:id.",
                "SuccessMs.jsx loads booking, self-drive vs chauffeur copy, links to KYC/track/account.",
                "Done",
                "P0",
                "",
            ),
            (
                "FE.W.08",
                "Track booking",
                "/track/:bookingId",
                "Timeline: HOLD → payment → KYC → signature → confirmed → handover → ongoing → return → completed. Logged-in users use GET /v1/bookings/:id. Logged-out: id + phone OTP. Show driver name after assignment.",
                "Track.jsx timeline, driver/vehicle, socket subscribe. Guest: publicId + phone OTP.",
                "Done",
                "P0",
                "OTP is emailed via notify template; SMS later.",
            ),
            (
                "FE.W.09",
                "Blogs",
                "/blogs  /blogs/:slug",
                "List published posts with hero image, date, category. Detail shows body + comments (if enabled). SEO title/description from CMS.",
                "AllBlogs / Blogspage ported (old context/Firestore). API GET /v1/public/blogs and /:slug exists.",
                "In Progress",
                "P1",
                "Point UI at platform API; drop Firestore blog context.",
            ),
            (
                "FE.W.10",
                "Contact / enquiry",
                "/contact",
                "Name, phone or email, city, message. Submit POST /v1/public/contact (creates CRM lead). Success toast. WhatsApp CTA can fire POST /v1/public/leads with source=whatsapp.",
                "Ported Contact.jsx + WhatsAppPopup. May still hit old backend.",
                "In Progress",
                "P1",
                "Switch to API; track WhatsApp clicks as leads.",
            ),
            (
                "FE.W.11",
                "CMS / legal pages",
                "/about  /howitworks  /termsandconditions  /legal/:slug",
                "Render published CmsPage by slug (title + body + meta). Admin can update without a deploy.",
                "Static About, HowItWorks, TermsAndConditions components. GET /v1/public/pages/:slug exists unused.",
                "In Progress",
                "P1",
                "Drive these from CMS; add /legal/:slug.",
            ),
            (
                "FE.W.12",
                "Packages (airport / tours)",
                "/packages  /packages/:slug",
                "List published tour packages with days, price, inclusions. Detail shows itinerary. CTA goes to /fleet with TOUR_PACKAGE + packageId + car class.",
                "Packages.jsx listing + detail. Header Tours link. CarDetail collects airport terminal/flight/wait, outstation km, one-way drop branch, tour package.",
                "Done",
                "P2",
                "",
            ),
            (
                "FE.W.13",
                "Header, footer, global chrome",
                "all public routes",
                "Nav: Home, Fleet, Packages, Blogs, Contact, Login/Account. After login show name + KYC badge. City selector. Footer links to legal CMS pages.",
                "Ported Header/Footer. AuthProvider wraps the app but header may not use it.",
                "In Progress",
                "P0",
                "Auth-aware nav; link Fleet to /fleet not old carousel.",
            ),
            (
                "FE.W.14",
                "Testimonials",
                "/testimonials  (and home module)",
                "Published testimonials on home. Optional listing page.",
                "Testimonial.jsx + TestimonialContext (old). No new API.",
                "Not Started",
                "P1",
                "Needs public testimonials API first, then wire home.",
            ),
        ],
    ),
    (
        "FE Web Account",
        "Frontend — customer dashboard (apps/web)",
        "apps/web :3000",
        "P0",
        "In Progress",
        "Logged-in area. Guard all /account/* with AuthContext; redirect to /login. Customer can only see own data. Support must not impersonate. Nine pages are listed starting at row 5 (scroll to the top of this tab).",
        [
            (
                "FE.A.01",
                "Account home",
                "/account",
                "Greeting, KYC status chip, shortcuts (bookings, KYC, invoices, tickets). Latest 5 bookings with status and track link. Edit name/phone.",
                "Account.jsx shows name, KYC, booking list, KYC link, sign out. GET /v1/me/bookings.",
                "In Progress",
                "P0",
                "Profile edit form (PATCH /v1/me); nav to invoices/wallet/tickets.",
            ),
            (
                "FE.A.02",
                "My bookings",
                "/account/bookings",
                "Table/cards of all bookings: publicId, car, dates, amount, status. Filter by status. Click → detail or /track/:id.",
                "List is inline on /account only. No dedicated route.",
                "In Progress",
                "P0",
                "Add /account/bookings page; empty state; cancel CTA with confirm.",
            ),
            (
                "FE.A.03",
                "Booking detail",
                "/account/bookings/:id",
                "Full timeline, vehicle/driver once assigned, pay remaining balance + deposit, links to KYC, agreement, invoices. Cancel until handover.",
                "GET /v1/bookings/:id includes history, payments, KYC, agreements, driver name/phone, inspections (photos/items/damages). Cancel until handover. Damage extras + inspection photos on the page.",
                "Done",
                "P0",
                "Pay remaining balance/deposit still on checkout/track.",
            ),
            (
                "FE.A.04",
                "KYC upload",
                "/account/kyc",
                "Upload Aadhaar, PAN, DL, selfie (pdf/jpg/png). Show status NOT_STARTED → SUBMITTED → APPROVED/REJECTED. Rejection reason + re-upload. After approve, prompt to sign agreement.",
                "Kyc.jsx: per-doc pickers (Aadhaar/PAN/DL/selfie/address), hashed numbers, DL expiry, rejection/re-upload. POST /v1/kyc/uploads (multipart) + POST /v1/kyc/submit + GET /v1/me/kyc.",
                "Done",
                "P0",
                "Direct Cloudinary signed upload needs CLOUDINARY_API_SECRET; server upload works now.",
            ),
            (
                "FE.A.05",
                "Agreements",
                "/account/agreements",
                "List agreements per booking. Download draft/signed PDF when ready. Status DRAFT/SENT/SIGNED/WAIVED.",
                "/account/agreements · AccountAgreements.jsx · GET /v1/me/documents · GET /v1/me/agreements/:id/pdf and /signed-pdf.",
                "Done",
                "P0",
                "Leegality invite/sign URL is on the envelope after send (module 10).",
            ),
            (
                "FE.A.06",
                "Invoices",
                "/account/invoices",
                "List tax invoices (number, date, amount, GST). Download/print. Link to booking.",
                "AccountInvoices.jsx on /account/invoices. GET /v1/me/invoices. Authenticated PDF download via GET /v1/me/invoices/:id/pdf.",
                "Done",
                "P0",
                "",
            ),
            (
                "FE.A.07",
                "Wallet",
                "/account/wallet",
                "Show balance (INR) and last 50 txns. Later: pay part of token from wallet.",
                "No page. GET /v1/me/wallet exists (balance 0, no debit at checkout).",
                "Not Started",
                "P2",
                "Read-only UI is enough until wallet pay is built.",
            ),
            (
                "FE.A.08",
                "Support tickets",
                "/account/tickets",
                "Open a ticket (subject, body, optional booking). List status. Reply thread. Cannot see internal notes.",
                "No page. POST /v1/tickets exists. No customer list/reply UI.",
                "Not Started",
                "P2",
                "Add list + thread when GET /v1/me/tickets exists (today only POST).",
            ),
            (
                "FE.A.09",
                "Addresses / profile",
                "/account (section) or /account/profile",
                "Edit full name, phone (re-OTP later), default city, addresses. Once KYC approved, name matches KYC unless admin override.",
                "PATCH /v1/me accepts fullName, phone, address JSON. No form.",
                "Not Started",
                "P0",
                "Simple profile form on account home is enough for MVP.",
            ),
        ],
    ),
    (
        "FE Admin",
        "Frontend — operations console (apps/admin :3001)",
        "apps/admin",
        "P0–P1",
        "In Progress",
        "Staff-only Next.js app. Token in localStorage (dd_token). Sidebar Shell. Role-aware nav later (SUPER_ADMIN sees all). Call the same API :4000. Old Firestore admin still lives under apps/web /admin — do not extend it; migrate screens here.",
        [
            (
                "FE.AD.01",
                "Admin login",
                "/login",
                "Staff email login (Firebase or dev token). Reject CUSTOMER role. Redirect to dashboard.",
                "apps/admin/app/login/page.jsx exists.",
                "In Progress",
                "P0",
                "Block non-staff; show role; production Firebase.",
            ),
            (
                "FE.AD.02",
                "Dashboard",
                "/",
                "KPI cards: bookings, pending KYC, pending signatures, failed payments, today’s handovers, overdue returns, vehicles in workshop, revenue. Filter city + date. Status breakdown table.",
                "app/page.jsx KPI cards including handovers, overdue returns, KYC, signatures, failed payments, workshop, free vehicles, bookings, revenue. Date filter + shell city/branch.",
                "Done",
                "P1",
                "",
            ),
            (
                "FE.AD.03",
                "Bookings list",
                "/bookings",
                "Filter by status, city, dates, q. Columns: publicId, customer, car, dates, amount, status. Row click → detail.",
                "List with status/q filters, create-on-behalf form, row links to detail. GET /v1/admin/bookings.",
                "Done",
                "P0",
                "",
            ),
            (
                "FE.AD.04",
                "Booking detail + ops",
                "/bookings/[id]",
                "Timeline, notes, flight number. Actions: change status, assign vehicle, assign driver, handover (odometer, fuel, photos), return (+ damages), generate/send agreement, mark signed, record offline pay. Show payments/KYC/inspections.",
                "apps/admin/app/bookings/[id]: timeline, re-quote dates, assign vehicle/driver, status+comped, policy cancel, handover/return inspections (odo, fuel, photos, accessories, damages). Socket subscribe.",
                "Done",
                "P0",
                "",
            ),
            (
                "FE.AD.05",
                "Car models",
                "/cars",
                "CRUD models: name, slug, type, seats, fuel, transmission, city, published, featured, images, pricing rules (daily/hourly/deposit).",
                "app/cars/page.jsx + GET/POST/PATCH /v1/admin/car-models and pricing-rules.",
                "In Progress",
                "P0",
                "Image manager; inline pricing editor.",
            ),
            (
                "FE.AD.06",
                "Vehicles",
                "/vehicles",
                "CRUD physical cars: registration, model, branch, status, odometer, partner. Queue or immediate branch transfer. RC/insurance/PUC/permit upload (local files if no Cloudinary) + expiry emails. Backfill missing docs. Partner contract required for partner cars.",
                "app/vehicles/page.jsx — Fleet / Expiries / Cities & branches. transfer-branch, vehicle-transfers complete/cancel, documents CRUD, expiries/notify, backfill-documents.",
                "Done",
                "P0",
                "",
            ),
            (
                "FE.AD.07",
                "Availability calendar",
                "/availability",
                "Per city/branch, see vehicles blocked by bookings, HOLD, maintenance. Staff can add a manual block.",
                "No page. Availability is computed in catalog search.",
                "Not Started",
                "P1",
                "Read-only calendar first; manual block later.",
            ),
            (
                "FE.AD.08",
                "Customers",
                "/customers",
                "Search users. View profile, roles, bookings. SUPER_ADMIN: set roles, disable. No impersonation.",
                "app/customers/page.jsx + GET /v1/admin/users, PATCH roles, PUT /v1/admin/staff/:id/scope, POST disable. Invite sets city/branch.",
                "In Progress",
                "P0",
                "Customer drawer with bookings + KYC.",
            ),
            (
                "FE.AD.09",
                "KYC queue",
                "/kyc",
                "Queue SUBMITTED/UNDER_REVIEW. Open docs. Approve/reject with notes. Decision advances self-drive booking.",
                "app/kyc/page.jsx: status filter, doc viewer, reject reason required, per-doc re-upload, UNDER_REVIEW, DL expiry flags. GET /v1/admin/kyc + POST decision/review/request-reupload.",
                "Done",
                "P0",
                "SALES is view-only; SUPPORT/SUPER_ADMIN decide.",
            ),
            (
                "FE.AD.10",
                "Agreements",
                "/agreements",
                "List envelopes. Generate PDF, send Leegality, mark signed (waiver), void/re-issue. Templates by city/product.",
                "app/agreements/page.jsx + Shell. GET /v1/admin/agreements, generate/send/mark-signed/void, template CRUD. Booking detail has Generate.",
                "Done",
                "P0",
                "Live Leegality still needs workflow credentials (module 10).",
            ),
            (
                "FE.AD.11",
                "Payments / invoices / deposits",
                "/payments  /invoices  /deposits",
                "List payments. Record offline cash/UPI. Refund (finance). Capture/release deposit. Show invoice numbers.",
                "Tabbed page: payments (refund/offline), invoices (GST + PDF), deposits (capture/release), recon (run Razorpay settlement match). GET /v1/admin/payments|invoices|deposits + POST /v1/admin/payments/reconcile.",
                "Done",
                "P0",
                "",
            ),
            (
                "FE.AD.12",
                "Drivers",
                "/drivers",
                "CRUD chauffeurs, branch, active flag. Documents (DL/badge), leave, current assignment. Availability tab for a date window.",
                "app/drivers roster + docs/leave + Availability tab. GET /v1/admin/drivers/availability. Assign on booking detail.",
                "Done",
                "P1",
                "",
            ),
            (
                "FE.AD.13",
                "Maintenance",
                "/maintenance",
                "Create workshop jobs with dates → blocks availability. Complete job returns vehicle AVAILABLE. Parts/cost later.",
                "No Shell link. POST /v1/admin/maintenance-jobs exists.",
                "Not Started",
                "P2",
                "Simple list + create is enough after booking path.",
            ),
            (
                "FE.AD.14",
                "Handover / inspections",
                "/inspections  or booking detail",
                "Start handover: odometer, fuel, photos. Start return: same + damage lines (amount). Completes booking and posts partner ledger.",
                "app/inspections + booking detail handover/return (odo, fuel, photos, accessories, signature, damage lines, close to release deposit).",
                "Done",
                "P1",
                "",
            ),
            (
                "FE.AD.15",
                "Partners + ledger + settlements",
                "/partners  /partners/[id]/ledger  /settlements",
                "CRUD partners, commission %. Ledger of trip earnings. Generate settlement period, mark paid with UTR. No partner login.",
                "app/partners/page.jsx list. Ledger/settlement APIs exist, no UI.",
                "In Progress",
                "P2",
                "Partner detail + generate/mark-paid.",
            ),
            (
                "FE.AD.16",
                "Cities, branches, staff",
                "/cities  /branches  /staff",
                "CRUD cities/branches. Assign StaffScope (user + city + branch). City/branch switcher in the shell.",
                "app/cities/page.jsx + app/staff/page.jsx. Shell city/branch switcher sends x-ops-city-id / x-ops-branch-id. PUT /v1/admin/staff/:id/scope.",
                "Done",
                "P1",
                "",
            ),
            (
                "FE.AD.17",
                "Offers",
                "/offers",
                "Create percent/flat codes with date window and max redemptions. Freeze abusive codes later.",
                "app/offers/page.jsx + GET/POST /v1/admin/offers.",
                "In Progress",
                "P1",
                "Disable/freeze; city/product limits when API supports them.",
            ),
            (
                "FE.AD.18",
                "CMS, banners, blogs, media",
                "/cms  /blogs  /banners  /media",
                "Edit pages by slug, banners (image, link, dates), blogs, media library. Publish/unpublish.",
                "app/cms/page.jsx + CMS/banner APIs. No blogs/media screens in admin app.",
                "In Progress",
                "P1",
                "Banners + blogs editors; do not grow apps/web/Admin/*.",
            ),
            (
                "FE.AD.19",
                "Leads CRM",
                "/leads",
                "Pipeline NEW → CONTACTED → QUALIFIED → BOOKED/LOST. Notes. Convert to quote later.",
                "app/leads/page.jsx + list/status/notes APIs. Convert not built.",
                "In Progress",
                "P1",
                "Kanban or status dropdown; convert when API exists.",
            ),
            (
                "FE.AD.20",
                "Tickets & reviews",
                "/tickets  /reviews",
                "Ticket queue: assign, reply, internal notes. Reviews: publish/hide.",
                "No Shell links. GET /v1/admin/tickets and /reviews exist. No PATCH moderate.",
                "Not Started",
                "P2",
                "List views first; moderate when PATCH exists.",
            ),
            (
                "FE.AD.21",
                "Reports",
                "/reports  /reports/revenue  /reports/gst",
                "Date-range revenue, bookings, GST. Export CSV later. Finance + SUPER_ADMIN.",
                "app/reports/page.jsx + GET /v1/admin/reports/:kind (revenue or gst).",
                "In Progress",
                "P2",
                "Date filters; CSV download.",
            ),
            (
                "FE.AD.22",
                "Notification log + templates",
                "/notifications",
                "Delivery log. Edit email templates by key. Resend later.",
                "No page. GET logs + PUT templates exist.",
                "Not Started",
                "P1",
                "Simple table + template editor.",
            ),
            (
                "FE.AD.23",
                "Trips, tours & airports",
                "/packages",
                "Tabs: tour package CRUD (days/inclusions/class), city-pair one-way fees, airport terminals (wait + night), outstation driver-allowance policy. Booking detail can add extra km/hours.",
                "app/packages/page.jsx + Shell Trips & tours. Booking extras on /bookings/[id].",
                "Done",
                "P1",
                "",
            ),
        ],
    ),
    (
        "FE Mobile",
        "Frontend — customer Expo app (apps/mobile)",
        "apps/mobile",
        "P2",
        "In Progress",
        "Same /v1 APIs as web. Do not build full navigation until web checkout + KYC + pay work. Current App.tsx is a single-screen prototype.",
        [
            (
                "FE.M.01",
                "Splash / force-update",
                "screen 1",
                "Brand splash. If app version < min, block with store link.",
                "Not built.",
                "Not Started",
                "P2",
                "Needs min-version from API or remote config.",
            ),
            (
                "FE.M.02",
                "Login / OTP",
                "screen 2",
                "Email OTP or Firebase phone/email. Store token. POST /v1/auth/sync. POST /v1/me/devices for FCM.",
                "Dev token only (Bearer dev:email) on App.tsx.",
                "Not Started",
                "P2",
                "No OTP UI.",
            ),
            (
                "FE.M.03",
                "Home + search",
                "screens 3–4",
                "City + dates + type. Results from GET /v1/public/search.",
                "App.tsx lists cars from search on load, no filters.",
                "In Progress",
                "P2",
                "Filters, pull-to-refresh, empty state.",
            ),
            (
                "FE.M.04",
                "Car detail",
                "screen 5",
                "Gallery, availability calendar, quote CTA.",
                "Not built (name + price only on list).",
                "Not Started",
                "P2",
                "",
            ),
            (
                "FE.M.05",
                "Checkout + Razorpay",
                "screens 6–7",
                "Quote → booking → Razorpay native checkout → verify/poll.",
                "Not built.",
                "Not Started",
                "P2",
                "After web Razorpay path is proven.",
            ),
            (
                "FE.M.06",
                "KYC + agreement",
                "screen 8",
                "Upload docs (camera/gallery). Open Leegality in-app browser.",
                "Not built.",
                "Not Started",
                "P2",
                "",
            ),
            (
                "FE.M.07",
                "Bookings + tracking",
                "screens 9–10",
                "List + timeline. Deep link dreamdrive://bookings/:id.",
                "App.tsx shows publicId + status after dev login.",
                "In Progress",
                "P2",
                "Detail timeline; push open.",
            ),
            (
                "FE.M.08",
                "Profile, tickets, notifications",
                "screens 11–13",
                "Profile, support ticket, in-app notification list.",
                "Not built.",
                "Not Started",
                "P2",
                "",
            ),
        ],
    ),
]


def style_header(ws, row=4, headers=None):
    headers = headers or HEADERS
    for col, title in enumerate(headers, 1):
        cell = ws.cell(row, col, title)
        cell.fill = FILLS["header"]
        cell.font = FONTS["header"]
        cell.alignment = Alignment(wrap_text=True, vertical="center", horizontal="center")
        cell.border = THIN
    ws.row_dimensions[row].height = 28
    last_col = get_column_letter(len(headers))
    ws.freeze_panes = f"A{row + 1}"


def add_status_validation(ws, first_data_row, last_data_row):
    dv = DataValidation(type="list", formula1='"Done,In Progress,Under Review,Not Started"', allow_blank=False)
    dv.error = "Pick Done, In Progress, Under Review, or Not Started"
    dv.errorTitle = "Invalid status"
    dv.prompt = "Status"
    dv.add(f"F{first_data_row}:F{last_data_row}")
    ws.add_data_validation(dv)


def add_status_cf(ws, first_data_row, last_data_row):
    col_f = f"$F{first_data_row}:$F{last_data_row}"
    mapping = [
        ('$F4="Done"', "C6EFCE", "006100"),
        ('$F4="In Progress"', "FFF2CC", "9C5700"),
        ('$F4="Under Review"', "FCE4D6", "C65911"),
        ('$F4="Not Started"', "F4CCCC", "9C0006"),
    ]
    # FormulaRule uses the first cell of the range
    for formula, bg, fg in [
        ('$F4="Done"', "C6EFCE", "006100"),
        ('$F4="In Progress"', "FFF2CC", "9C5700"),
        ('$F4="Under Review"', "FCE4D6", "C65911"),
        ('$F4="Not Started"', "F4CCCC", "9C0006"),
    ]:
        ws.conditional_formatting.add(
            f"A{first_data_row}:H{last_data_row}",
            FormulaRule(
                formula=[formula.replace("4", str(first_data_row))],
                fill=PatternFill("solid", fgColor=bg),
            ),
        )


def write_module_sheet(wb, tab, title, service, priority, overall, notes, rows, headers=None):
    headers = headers or HEADERS
    ws = wb.create_sheet(tab[:31])
    ws.sheet_properties.tabColor = {
        "Done": "006100",
        "In Progress": "F4B942",
        "Under Review": "E07A3D",
        "Not Started": "C1121F",
    }.get(overall, "40916C")

    ws.merge_cells("A1:H1")
    ws["A1"] = title
    ws["A1"].font = FONTS["title"]
    ws.merge_cells("A2:H2")
    ws["A2"] = f"Service: {service}    |    Priority: {priority}    |    Overall status: {overall}"
    ws["A2"].font = FONTS["status"][overall]
    ws.merge_cells("A3:H3")
    ws["A3"] = notes
    ws["A3"].font = FONTS["sub"]
    ws["A3"].alignment = WRAP
    ws.row_dimensions[3].height = 48

    style_header(ws, 4, headers)
    for i, row in enumerate(rows, 5):
        for c, value in enumerate(row, 1):
            cell = ws.cell(i, c, value)
            cell.font = FONTS["body"]
            cell.alignment = WRAP
            cell.border = THIN
            if c == 6:
                cell.font = FONTS["status"].get(value, FONTS["bold"])
                cell.fill = FILLS.get(value, FILLS["alt"])
                cell.alignment = Alignment(horizontal="center", vertical="center")
            elif i % 2 == 0 and c != 6:
                cell.fill = FILLS["alt"]
        ws.row_dimensions[i].height = 56 if headers is not HEADERS else 42

    last = 4 + len(rows)
    add_status_validation(ws, 5, last)
    ws.column_dimensions["A"].width = 10
    ws.column_dimensions["B"].width = 36 if headers is HEADERS else 28
    ws.column_dimensions["C"].width = 18 if headers is HEADERS else 38
    ws.column_dimensions["D"].width = 62
    ws.column_dimensions["E"].width = 48
    ws.column_dimensions["F"].width = 16
    ws.column_dimensions["G"].width = 12
    ws.column_dimensions["H"].width = 48
    ws.page_setup.orientation = "landscape"
    ws.page_setup.fitToPage = True
    ws.page_setup.fitToWidth = 1
    ws.page_setup.fitToHeight = 0
    ws.sheet_view.showGridLines = False
    ws.auto_filter.ref = f"A4:{get_column_letter(len(headers))}{last}"

    # counts footer
    r = last + 2
    ws.cell(r, 1, "Counts").font = FONTS["bold"]
    labels = ["Done", "In Progress", "Under Review", "Not Started"]
    for i, lab in enumerate(labels):
        ws.cell(r, 2 + i, lab).font = FONTS["status"][lab]
        c = ws.cell(r + 1, 2 + i, f'=COUNTIF(F5:F{last},"{lab}")')
        c.font = FONTS["bold"]
        c.fill = FILLS[lab]
        c.alignment = Alignment(horizontal="center")
    return ws.title, overall, priority, service, notes, len(rows)


def write_legend(wb):
    ws = wb.create_sheet("Legend", 0)
    ws.sheet_properties.tabColor = "1B4332"
    ws.merge_cells("A1:F1")
    ws["A1"] = "Dream-Drive MS — functionality tracker"
    ws["A1"].font = FONTS["title"]
    ws.merge_cells("A2:F2")
    ws["A2"] = (
        "Greenfield rebuild at dream-drive-MS (NestJS API :4000 + Next.js web/admin + Expo). "
        "Live product still in client-main/backend. Snapshot date: 2026-08-31."
    )
    ws["A2"].alignment = WRAP
    ws.row_dimensions[2].height = 36

    ws["A4"] = "How to read status"
    ws["A4"].font = FONTS["bold"]
    legend = [
        ("Done", "Coded and usable (API + persistence). May still need production keys or UI polish."),
        ("In Progress", "Partial: endpoint or screen exists but validations, UI, or edge cases are incomplete."),
        ("Under Review", "Implemented with mock/fallback (Razorpay, Leegality, Gmail, Firebase). Needs staging verification."),
        ("Not Started", "Not built in this repo yet (or explicitly out of scope)."),
    ]
    ws["A5"] = "Status"
    ws["B5"] = "Meaning"
    ws["A5"].font = FONTS["header"]
    ws["B5"].font = FONTS["header"]
    ws["A5"].fill = FILLS["header"]
    ws["B5"].fill = FILLS["header"]
    for i, (st, meaning) in enumerate(legend, 6):
        ws.cell(i, 1, st).fill = FILLS[st]
        ws.cell(i, 1).font = FONTS["status"][st]
        ws.cell(i, 2, meaning).alignment = WRAP
        ws.row_dimensions[i].height = 28

    ws["A11"] = "Out of scope (confirmed)"
    ws["A11"].font = FONTS["bold"]
    ws["A12"] = "Delivery / pickup executive operations"
    ws["A13"] = "Partner self-service portal (admin manages partners)"
    ws["A15"] = "Stack"
    ws["A15"].font = FONTS["bold"]
    ws.merge_cells("A16:F16")
    ws["A16"] = (
        "Next.js 15 web :3000 (ported marketing site + new /fleet checkout path) · "
        "Next.js admin :3001 · NestJS API :4000 (identity, catalog, booking, payment, document, fleet, partner, notify, platform) · "
        "worker HOLD sweeper · socket :4010 · PostgreSQL/Prisma · Firebase Auth · Razorpay · Leegality · Zoho Forms · Expo mobile."
    )
    ws["A16"].alignment = WRAP
    ws.row_dimensions[16].height = 48

    ws["A18"] = "How to refresh this Google Sheet"
    ws["A18"].font = FONTS["bold"]
    ws.merge_cells("A19:F19")
    ws["A19"] = (
        "File → Import → Upload Dream-Drive-Functionality-Tracker.xlsx → "
        "Import location: Replace spreadsheet. Each module is its own tab. Change Status via the dropdown; colours follow the cell."
    )
    ws["A19"].alignment = WRAP
    ws.row_dimensions[19].height = 36

    ws["A21"] = "Frontend developer tabs"
    ws["A21"].font = FONTS["bold"]
    ws.merge_cells("A22:F22")
    ws["A22"] = (
        "FE Web Public, FE Web Account, FE Admin, FE Mobile — one row per page: route, how it should function, "
        "what exists today, APIs, status. Same dropdown: Done / In Progress / Under Review / Not Started."
    )
    ws["A22"].alignment = WRAP
    ws.row_dimensions[22].height = 36

    ws.column_dimensions["A"].width = 18
    ws.column_dimensions["B"].width = 100
    ws.sheet_view.showGridLines = False


def write_summary(wb, module_meta, index=1):
    ws = wb.create_sheet("00 Summary", index)
    ws.sheet_properties.tabColor = "2D6A4F"
    ws.merge_cells("A1:L1")
    ws["A1"] = "Dream-Drive MS — module + frontend summary"
    ws["A1"].font = FONTS["title"]
    ws.merge_cells("A2:L2")
    ws["A2"] = "Change Status on each module tab. Counts below use COUNTIF against that tab. Overall status is the current engineering judgement for the module as a whole."
    ws["A2"].font = FONTS["sub"]
    ws["A2"].alignment = WRAP

    headers = [
        "#",
        "Module",
        "Service",
        "Priority",
        "Overall status",
        "Done",
        "In Progress",
        "Under Review",
        "Not Started",
        "Total items",
        "% Done",
        "Notes",
    ]
    for col, h in enumerate(headers, 1):
        cell = ws.cell(4, col, h)
        cell.fill = FILLS["header"]
        cell.font = FONTS["header"]
        cell.alignment = Alignment(wrap_text=True, vertical="center", horizontal="center")
        cell.border = THIN
    ws.freeze_panes = "A5"
    ws.auto_filter.ref = "A4:L4"

    for i, (tab, overall, priority, service, notes, n) in enumerate(module_meta, 5):
        # quoted sheet names
        q = f"'{tab}'"
        ws.cell(i, 1, i - 4)
        ws.cell(i, 2, tab)
        ws.cell(i, 3, service)
        ws.cell(i, 4, priority)
        st = ws.cell(i, 5, overall)
        st.fill = FILLS[overall]
        st.font = FONTS["status"][overall]
        st.alignment = Alignment(horizontal="center")
        ws.cell(i, 6, f"=COUNTIF({q}!F:F,\"Done\")")
        ws.cell(i, 7, f"=COUNTIF({q}!F:F,\"In Progress\")")
        ws.cell(i, 8, f"=COUNTIF({q}!F:F,\"Under Review\")")
        ws.cell(i, 9, f"=COUNTIF({q}!F:F,\"Not Started\")")
        ws.cell(i, 10, f"=F{i}+G{i}+H{i}+I{i}")
        ws.cell(i, 11, f"=IF(J{i}=0,0,F{i}/J{i})")
        ws.cell(i, 11).number_format = "0%"
        ws.cell(i, 12, notes).alignment = WRAP
        for c in range(1, 13):
            ws.cell(i, c).border = THIN
            if c != 5:
                ws.cell(i, c).font = FONTS["body"]
            if i % 2 == 0 and c not in (5,):
                ws.cell(i, c).fill = FILLS["alt"]
        ws.row_dimensions[i].height = 48

    last = 4 + len(module_meta)
    tot = last + 1
    ws.cell(tot, 2, "TOTAL").font = FONTS["bold"]
    for col, letter in enumerate(["F", "G", "H", "I", "J"], 6):
        cell = ws.cell(tot, col, f"=SUM({letter}5:{letter}{last})")
        cell.font = FONTS["bold"]
        cell.fill = FILLS["card"]
    ws.cell(tot, 11, f"=IF(J{tot}=0,0,F{tot}/J{tot})").number_format = "0%"
    ws.cell(tot, 11).font = FONTS["bold"]

    # status rollup for chart (place after totals so it does not overlap module rows)
    rollup = tot + 2
    ws.cell(rollup, 1, "Status rollup (all items)").font = FONTS["bold"]
    ws.cell(rollup + 1, 1, "Status").fill = FILLS["header"]
    ws.cell(rollup + 1, 2, "Count").fill = FILLS["header"]
    ws.cell(rollup + 1, 1).font = FONTS["header"]
    ws.cell(rollup + 1, 2).font = FONTS["header"]
    for i, st in enumerate(STATUSES):
        r = rollup + 2 + i
        ws.cell(r, 1, st).fill = FILLS[st]
        ws.cell(r, 1).font = FONTS["status"][st]
        col = ["F", "G", "H", "I"][i]
        ws.cell(r, 2, f"={col}{tot}")

    pie = PieChart()
    pie.title = "Items by status"
    labels = Reference(ws, min_col=1, min_row=rollup + 2, max_row=rollup + 5)
    data = Reference(ws, min_col=2, min_row=rollup + 1, max_row=rollup + 5)
    pie.add_data(data, titles_from_data=True)
    pie.set_categories(labels)
    pie.dataLabels = DataLabelList()
    pie.dataLabels.showPercent = True
    pie.dataLabels.showVal = False
    pie.dataLabels.showCatName = False
    pie.width = 12
    pie.height = 8
    ws.add_chart(pie, f"D{rollup}")

    widths = [6, 24, 28, 10, 16, 12, 14, 14, 14, 12, 12, 70]
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w
    ws.row_dimensions[4].height = 24
    ws.sheet_view.showGridLines = False
    dv = DataValidation(type="list", formula1='"Done,In Progress,Under Review,Not Started"', allow_blank=False)
    dv.add(f"E5:E{last}")
    ws.add_data_validation(dv)


def main():
    wb = Workbook()
    default = wb.active
    wb.remove(default)

    write_legend(wb)
    meta = []
    fe_order = ["FE Web Account", "FE Web Public", "FE Admin", "FE Mobile"]
    fe_by_name = {m[0]: m for m in FE_MODULES}
    for key in fe_order:
        tab, title, service, priority, overall, notes, rows = fe_by_name[key]
        name, overall, priority, service, notes, n = write_module_sheet(
            wb, tab, title, service, priority, overall, notes, rows, headers=FE_HEADERS
        )
        meta.append((name, overall, priority, service, notes, n))
    for tab, title, service, priority, overall, notes, rows in MODULES:
        name, overall, priority, service, notes, n = write_module_sheet(
            wb, tab, title, service, priority, overall, notes, rows
        )
        meta.append((name, overall, priority, service, notes, n))
    write_summary(wb, meta, index=1 + len(fe_order))

    # move summary next to legend: already inserted at index 1
    OUT.parent.mkdir(parents=True, exist_ok=True)
    wb.save(OUT)
    print(f"Wrote {OUT}")
    print(f"Modules: {len(meta)}  Items: {sum(m[-1] for m in meta)}")


if __name__ == "__main__":
    main()
